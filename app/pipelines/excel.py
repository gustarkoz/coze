import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from app.render import render_html

def _sheet_name(raw: str, used: set) -> str:
    name = re.sub(r'[\\/*\[\]:?]', '', raw).strip()
    if not name:
        name = 'Sheet'
    name = name[:31]
    base = name
    i = 2
    while name in used:
        name = (base[:28] + f"_{i}")[:31]
        i += 1
    used.add(name)
    return name

def markdown_to_xlsx(md_text: str, output_path: str) -> list:
    html = render_html(md_text, 'Document')
    soup = BeautifulSoup(html, 'html.parser')
    tables = soup.find_all('table')
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    sheets = []
    idx = 1
    for t in tables:
        heading = t.find_previous(['h1', 'h2', 'h3'])
        title = heading.get_text(strip=True) if heading else f'Table {idx}'
        idx += 1
        sn = _sheet_name(title, used)
        ws = wb.create_sheet(sn)
        sheets.append(sn)
        rows = t.find_all('tr')
        for r in rows:
            cells = r.find_all(['th', 'td'])
            ws.append([c.get_text(strip=True) for c in cells])
    if not tables:
        ws = wb.create_sheet('Sheet1')
        ws.append(['No tables detected'])
        sheets.append('Sheet1')
    wb.save(output_path)
    return sheets
